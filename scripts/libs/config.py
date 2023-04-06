from typing import Union
import sys
import os
import openpyxl

from pldata import ProfitDataItem, LossDataItem, LabelManager
from excel import utils


def read_config_file(file_path: str, data_store: Union[dict, None], label_mgr: LabelManager):
    if not os.path.exists(file_path):
        print(f"XXX {file_path}が存在しません。ディレクトリを確認してください")
        sys.exit(1)
    filename = os.path.splitext(os.path.basename(file_path))[0]
    print(f"* reading: {filename}.xlsx")
    workbook = openpyxl.load_workbook(file_path)
    _parse_config(data_store, label_mgr, workbook)


def _parse_config(store: dict, label_mgr: LabelManager, wb: any):
    for ws_name in wb.sheetnames:
        if ws_name == "設定":
            store.setdefault("config", {})
            _read_misc_config(store["config"], wb, ws_name)
        else:
            input_data = store.setdefault("definition", {}).setdefault(ws_name, {})  # ファイルを読み込むたびにdefinitionは刷新する（古い設定は消してから作り直す）
            input_data.setdefault("profit", [])
            input_data.setdefault("loss", [])
            _read_pl_items(input_data, label_mgr, wb, ws_name)


def _read_misc_config(conf: dict, wb: any, ws_name: str):
    """各種全体設定を読み込む"""
    ws = wb[ws_name]
    # 売上に関する設定と、経費に関する設定がそれぞれ何カラム目から始まっているかを見つける
    r = utils.find_column_numbers(["設定項目"], ws)
    if "設定項目" not in r:
        print(f"XXX {ws_name}は不正なシートです")
        return
    conf_col = r["設定項目"][0]       # type: int
    conf_row = r["設定項目"][1]       # type: int

    for row in ws.iter_rows(min_row=conf_row+1):
        if conf_col > -1 and row[conf_col].value is not None:
            if row[conf_col+1].value is not None and row[conf_col+1].value != "":
                conf[row[conf_col].value] = row[conf_col+1].value


def _read_pl_items(input_data: dict, label_mgr: LabelManager, wb: any, business: str):
    """売上項目と経費項目を列挙したシートを読み込む"""
    ws = wb[business]
    # 売上に関する設定と、経費に関する設定がそれぞれ何カラム目から始まっているかを見つける
    r = utils.find_column_numbers(["売上項目", "経費グループ"], ws)
    if "売上項目" not in r or "経費グループ" not in r:
        print(f"XXX 設定.xlsxの{business}は不正なシートです")
        return

    sales_col = r["売上項目"][0]       # type: int
    expense_col = r["経費グループ"][0]  # type: int

    # １行目にタイトルラベルが入っている前提で考える
    # 一行ずつ設定を読み込む(最初の行はヘッダ行なのでスキップする)
    for row in ws.iter_rows(min_row=2):
        if sales_col > -1 and row[sales_col].value is not None:
            # 売上に関する設定情報を読み込む(すでに同じ名称があれば上書きする)
            idx = next((i for i, d in enumerate(input_data["profit"]) if d.name == row[sales_col].value), None)
            item = ProfitDataItem(row[sales_col].value, row[sales_col+1].value)
            if idx is None:
                input_data["profit"].append(item)
            else:
                input_data["profit"][idx] = item
            label_mgr.add(business, "profit", item)

        if expense_col > -1 and row[expense_col].value is not None:
            # 経費に関する設定情報を読み込む(すでに同じ名称があれば上書きする)
            idx = next((i for i, d in enumerate(input_data["loss"]) if d.group == row[expense_col].value and d.account == row[expense_col+1].value and d.category == row[expense_col+2].value), None)
            item = LossDataItem(row[expense_col].value, row[expense_col+1].value, row[expense_col+2].value, row[expense_col+3].value, row[expense_col+4].value, row[expense_col+5].value)
            if idx is None:
                input_data["loss"].append(item)
            else:
                input_data["loss"][idx] = item
            label_mgr.add(business, "loss", item)

    if len(input_data["profit"]) == 0:
        input_data["profit"].append(ProfitDataItem(None, None))
    if len(input_data["loss"]) == 0:
        input_data["loss"].append(LossDataItem(None, None, None, None, None, None))
