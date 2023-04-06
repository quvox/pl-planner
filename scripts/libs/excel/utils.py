from typing import Tuple, Dict, Union

import openpyxl
import unicodedata
from openpyxl.styles import NamedStyle

from . import styles


def create_new_workbook() -> any:
    """新しいExcelワークブックを作成してスタイルを適用する"""
    wb = openpyxl.Workbook()
    wb.add_named_style(styles.header_date_style)
    wb.add_named_style(styles.column_label_style)
    return wb


def set_style_and_value(cell: any, value: any, style: Union[Dict[str, any], None] = None):
    set_style(cell, style)
    cell.value = value


def set_style(cell: any, style: Union[Dict[str, any], None] = None):
    if style is not None:
        if "style" in style:
            cell.style = style["style"]
        if "border" in style:
            cell.border = style["border"]
        if "format" in style:
            cell.number_format = style["format"]


def merge_cells(ws: any, start_pos: Tuple[int, int], end_pos: Tuple[int, int]):
    """start_pos=(x1, y1), end_pos=(x2, y2)の範囲のセルを結合する"""
    ws.merge_cells(start_row=start_pos[1], start_column=start_pos[0],
                   end_row=end_pos[1], end_column=end_pos[0])


def get_cell_coordinate(ws: any, row: int, column: int):
    cell = ws.cell(row=row, column=column)
    return cell.coordinate


def auto_adjust_column_width(ws: any):
    """シート内の全ての列の幅を自動調整する"""
    target_row = 0
    for col in ws.columns:
        max_length = 0
        while not hasattr(col[target_row], "column_letter"):
            target_row += 1
        column = col[target_row].column_letter  # A,B,Cなどの列名

        for cell in col:
            if cell.value is None: continue
            if isinstance(cell.value, int) or isinstance(cell.value, float):
                val = int(cell.value*100)/100   # 有効数字小数第2位まで
                l = _count_str_length(str(val))
            else:
                l = _count_str_length(cell.value)
            if l > max_length:
                max_length = l

        new_width = (max_length + 2) * 1.3
        ws.column_dimensions[column].width = new_width


def set_styles_to_column(ws: any, column: int, start_row: int, end_row: int, style: NamedStyle, border=None):
    for i in range(start_row, end_row):
        cell = ws.cell(row=i, column=column)
        cell.style = style
        cell.number_format = styles.number_format
        if border is not None:
            cell.border = border


def _count_str_length(text):
    text_counter = 0
    for c in text:
        j = unicodedata.east_asian_width(c)
        if 'F' == j:
            text_counter += 2
        elif 'H' == j:
            text_counter += 1
        elif 'W' == j:
            text_counter += 2
        elif 'Na' == j:
            text_counter += 1
        elif 'A' == j:
            text_counter += 2
        else:
            text_counter += 1

    return text_counter


def find_column_numbers(labels: list, ws: any):
    """与えられたラベルが何行目、何列目にあるか（A1から調べて最初に見つかる場所）を見つける"""
    result = dict()
    for i, col in enumerate(ws.columns):
        for k, cell in enumerate(col):
            for label in labels:
                if cell.value == label and label not in result:
                    result[label] = (i, k)  # 列, 行
            if len(result.keys()) == len(labels):
                break
    return result

